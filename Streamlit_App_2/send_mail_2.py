import openpyxl, smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os

def app():
    wb = openpyxl.load_workbook('MasterSheet_1.xlsx')
    sheet = wb['Sheet1']

    sender_email=os.getenv("EMAIL")
    password_email=os.getenv("PASSWORD")

    employees = {}
    manager = {}

    for r in range(2, sheet.max_row + 1):
        designation = sheet.cell(row=r, column=2).value
        taskStatus = sheet.cell(row=r, column=7).value
        if designation == 'Employee' and (taskStatus == 'DELAYED' or taskStatus == 'INPROCESS'):
            email = sheet.cell(row=r, column=9).value
            projectId = sheet.cell(row=r, column=1).value
            taskId = sheet.cell(row=r, column=4).value
            empId = sheet.cell(row=r, column=3).value
            taskStatus = sheet.cell(row=r,column=7).value

            dueDate_cell = sheet.cell(row=r, column=6)
            dueDate_value = dueDate_cell.value
            dueDate = datetime.strptime(str(dueDate_value), '%Y-%m-%d %H:%M:%S')
            days_remaining = (dueDate - datetime.now()).days + 1

            countCell = sheet.cell(row=r, column=8)
            mailCount = countCell.value
            countCell.value = mailCount + 1
            wb.save('MasterSheet_1.xlsx')

            if days_remaining < 0 and taskStatus == 'INPROCESS':
                sheet.cell(row=r, column=7).value = 'DELAYED'
                taskStatus = sheet.cell(row=r, column=7).value
                wb.save('MasterSheet_1.xlsx')

            if email not in employees:
                employees[email] = []

            employees[email].append({
                'empId': empId,
                'projectId': projectId,
                'taskId': taskId,
                'reminders': countCell.value,
                'taskStatus': taskStatus,
                'dueDate': dueDate,
                'days_remaining': days_remaining
            })

        elif designation == 'Manager':
            email = sheet.cell(row=r, column=9).value
            projectId = sheet.cell(row=r, column=1).value

            if email not in manager:
                manager[email] = []
            manager[email].append(projectId)

    # Print the result
    # for employee_mail, employee_info_list in employees.items():
    #     print(f'Employee Email: {employee_mail}')
    #     for emp_info in employee_info_list:
    #         empId = emp_info['empId']
    #         dueDate = emp_info['dueDate']
    #         taskStatus = emp_info['taskStatus']
    #         taskId = emp_info['taskId']
    #         reminders = emp_info['reminders']
    #         projectId = emp_info['projectId']
    #         print(
    #             f'  EmpID: {empId}, Due Date: {dueDate}, TaskID: {taskId}, Reminders: {reminders}, ProjectID: {projectId}, Task Status: {taskStatus}')

    # Create a dictionary to store managers and their employees
    managers_and_employees = {}

    # Populate the dictionary based on the existing data
    for employee_mail, employee_info_list in employees.items():
        for emp_info in employee_info_list:
            projectId = emp_info['projectId']

            for email, proj_Id in manager.items():
                if projectId in proj_Id:
                    manager_email = email

                    if manager_email not in managers_and_employees:
                        managers_and_employees[manager_email] = []

                    managers_and_employees[manager_email].append({
                        'empId': emp_info['empId'],
                        'dueDate': emp_info['dueDate'],
                        'taskId': emp_info['taskId'],
                        'reminders': emp_info['reminders'],
                        'projectId': emp_info['projectId'],
                        'taskStatus': emp_info['taskStatus'],
                        'days_remaining': emp_info['days_remaining']
                    })
                    break

    # Print the result
    # for manager_email, employees_list in managers_and_employees.items():
    #     print(f'Manager Email: {manager_email}')
    #     for emp_info in employees_list:
    #         print(f'EmpID: {emp_info["empId"]}, '
    #               f'Due Date: {emp_info["dueDate"]}, '
    #               f'TaskID: {emp_info["taskId"]}, '
    #               f'Reminders: {emp_info["reminders"]}, '
    #               f'ProjectID: {emp_info["projectId"]}, '
    #               f'Task Status: {emp_info["taskStatus"]}, '
    #               f'Days Remaining: {emp_info["days_remaining"]}')
    #     print()

    smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login('shreyag0104@outlook.com', 'gmvyylyowrootmis')

    # Send email to employees
    for employee_mail, employee_info_list in employees.items():
        employee_count = len(employee_info_list)
        empId = 'abc'
        dueDate = 'abc'
        taskId = 'abc'
        reminders = 0
        projectId = 'abc'
        days_remaining = 0
        verb = 'are'
        if employee_count == 1:
            verb = 'is'

        subject = "Task Reminder"
        body = f"This is just a quick note to remind you that {employee_count} out of your total tasks {verb} remaining.\n\n"

        for emp_info in employee_info_list:
            empId = emp_info['empId']
            dueDate = emp_info['dueDate']
            taskId = emp_info['taskId']
            reminders = emp_info['reminders']
            projectId = emp_info['projectId']
            taskStatus = emp_info['taskStatus']
            days_remaining = emp_info['days_remaining']

            # Add information to the body
            if days_remaining < 0:
                days_remaining = 0
            body += f" Project ID: {projectId}\n Task ID: {taskId}\n Due Date: {dueDate}\n Task Status: {taskStatus}\n Reminder No: {reminders}\n Days Remaining: {days_remaining}\n\n"

        body += f"\nBest Regards,\nAtlas Copco"
        # print(body)

        # Create MIMEText object
        msg = MIMEMultipart()
        msg.attach(MIMEText(body, 'plain'))

        # Set email headers
        msg['From'] = 'ATLAS COPCO <shreyag0104@outlook.com>'
        msg['To'] = employee_mail
        msg['Subject'] = subject

        # Send the email using smtplib
        sendmailStatus = smtpObj.sendmail('shreyag0104@outlook.com', employee_mail, msg.as_string())
        print(f'Sending email to {employee_mail}...')
        if sendmailStatus != {}:
            print('There was a problem sending email to %s: %s' % (employee_mail, sendmailStatus))

    # Send email to manager
    for manager_email, employees_list in managers_and_employees.items():
        # print(f'Manager Email: {manager_email}')
        employee_count = len(employees_list)
        empId = 'abc'
        dueDate = 'abc'
        taskId = 'abc'
        reminders = 0
        projectId = 'abc'
        verb = 'are'
        if employee_count == 1:
            verb = 'is'

        subject = "Weekly Update"
        body = f"This is just a quick note to update you about the remaining tasks.\n\n"

        for emp_info in employees_list:
            empId = emp_info['empId']
            dueDate = emp_info['dueDate']
            taskId = emp_info['taskId']
            reminders = emp_info['reminders']
            projectId = emp_info['projectId']
            taskStatus = emp_info['taskStatus']

            body += f" Project ID: {projectId}\n Task ID: {taskId}\n Task Status: {taskStatus}\n Due Date: {dueDate}\n Reminders Sent: {reminders}\n\n"

        body += f"\nBest Regards,\nAtlas Copco"
        # print(body)

        # Create MIMEText object
        msg = MIMEMultipart()
        msg.attach(MIMEText(body, 'plain'))

        # Set email headers
        msg['From'] = 'ATLAS COPCO <shreyag0104@outlook.com>'
        msg['To'] = manager_email
        msg['Subject'] = subject

        # Send the email using smtplib
        sendmailStatus = smtpObj.sendmail('shreyag0104@outlook.com', manager_email, msg.as_string())
        print(f'Sending email to {manager_email}...')
        if sendmailStatus != {}:
            print('There was a problem sending email to %s: %s' % (manager_email, sendmailStatus))

